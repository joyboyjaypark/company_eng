import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import webbrowser
import requests
import xml.etree.ElementTree as ET
from datetime import datetime
import tkinter.font as tkfont
import re
import threading

BASE_URL = "http://apis.data.go.kr/1360000/AsosHourlyInfoService/getWthrDataList"


class AsosGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("국내 지역별 온습도 엔탈피 검색기")
        self.geometry("900x700")
        self.resizable(True, True)

        self.create_widgets()

    def create_widgets(self):
        notebook = ttk.Notebook(self)
        notebook.pack(fill=tk.BOTH, expand=True)

        main_frame = ttk.Frame(notebook, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        notebook.add(main_frame, text='조회')

        summary_frame = ttk.Frame(notebook, padding=10)
        notebook.add(summary_frame, text='요약')

        # shared font (used by version and info labels)
        self.summary_font = tkfont.nametofont("TkDefaultFont")

        # Version tab
        version_frame = ttk.Frame(notebook, padding=10)
        notebook.add(version_frame, text='버전')

        # Key-in area at top
        key_frame = ttk.Frame(version_frame)
        key_frame.pack(fill=tk.X, anchor='n', pady=(0, 6))
        ttk.Label(key_frame, text="Key-in", font=self.summary_font).grid(row=0, column=0, sticky='w')
        ttk.Label(key_frame, text="입사연월일에 생년월일을 연속 16자리를 기입해주세요.", foreground='gray').grid(row=1, column=0, sticky='w')
        self.keyin_var = tk.StringVar()
        key_entry = ttk.Entry(key_frame, textvariable=self.keyin_var, width=40)
        key_entry.grid(row=0, column=1, rowspan=2, padx=(10,0))

        # Version info title and main text area (initially hidden until key matches)
        version_title = ttk.Label(version_frame, text="버전정보", font=self.summary_font)
        version_title.pack_forget()
        version_text = tk.Text(version_frame, height=10, width=80, wrap='word')
        version_text.pack_forget()
        version_text.insert(tk.END, """
1. 데이터 출처 : www.data.go.kr(공공데이터포털)
                     data.kma.go.kr(기상청 기사장료개방포털)
2. 상세 출처 : 기상청 지상(종관, ASOS) 시간자료 조회서비스
3. 일반인증키(Encoding) : Nv0jBnCHJXCT20iu910K%2FIGnF556Vt2w06icWR2uj66dF73AiTNBXaM7bIS9Nu9C0cmB7sGVgpnbCiK01Qkgeg%3D%3D
4. Version
   [1.0]
    - Excel -> Python으로 전환
    - Excel사용시 속도가 너무 느리고, 데이터 다운로드 페이지수가 100으로 한정되어 연단위로 받아야 하는 문제가 있었음.
   [개선필요]
    - TAC 온습도 산정시, 데이터 추출 시작일시와 범위를 명확히할 필요가 있음.
    - 윈도우상에 지도를 표기하여, 사용자가 주소를 입력하면 직선거리상으로 가장 가까운 기상관측소 3곳을 주소와의 직선거리로 표기하여 데이터 선택에 객관성을 추구하기 위함.
""")
        version_text.config(state='disabled')

        # show/hide logic based on exact 16-digit key
        def _on_keyin_change(*args):
            v = (self.keyin_var.get() or "").strip()
            if v == "2010010119840215":
                try:
                    version_title.pack(fill=tk.X, anchor='n', pady=(6,2))
                    version_text.pack(fill=tk.BOTH, expand=True)
                except Exception:
                    pass
            else:
                try:
                    version_text.pack_forget()
                    version_title.pack_forget()
                except Exception:
                    pass

        # bind trace
        try:
            self.keyin_var.trace_add('write', _on_keyin_change)
        except Exception:
            # fallback for older tkinter versions
            self.keyin_var.trace('w', lambda *a: _on_keyin_change())

        # top-right info label
        info_text = "제작일 : 25.11.28, 제작자 : 박재영, 버전 : 1.0"
        info_label = ttk.Label(main_frame, text=info_text, font=self.summary_font)
        info_label.grid(row=0, column=3, sticky='e')

        # layout
        for i in range(4):
            main_frame.columnconfigure(i, weight=1)

        row = 0
        ttk.Label(main_frame, text="1. 인증키", anchor="w").grid(row=row, column=0, sticky="w")
        row += 1
        self.service_key_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.service_key_var, width=80).grid(row=row, column=0, columnspan=4, sticky="we", pady=(0, 5))
        row += 1

        ttk.Label(main_frame, text="2. 응답자료형식 (dataType)", anchor="w").grid(row=row, column=0, sticky="w")
        self.data_type_var = tk.StringVar(value="JSON")
        data_type_combo = ttk.Combobox(main_frame, textvariable=self.data_type_var, values=["XML", "JSON"], width=8, state="readonly")
        data_type_combo.grid(row=row, column=1, sticky="w")
        row += 1

        ttk.Label(main_frame, text="3. 자료 코드 (dataCd) / 날짜 코드 (dateCd)", anchor="w").grid(row=row, column=0, sticky="w", pady=(5, 0))
        row += 1
        fixed_frame = ttk.Frame(main_frame)
        fixed_frame.grid(row=row, column=0, columnspan=4, sticky="w")
        ttk.Label(fixed_frame, text="dataCd (항상 ASOS): ").grid(row=0, column=0, sticky="w")
        self.data_cd_var = tk.StringVar(value="ASOS")
        ttk.Entry(fixed_frame, textvariable=self.data_cd_var, width=8, state="readonly").grid(row=0, column=1, sticky="w", padx=(0, 10))
        ttk.Label(fixed_frame, text="dateCd (항상 HR): ").grid(row=0, column=2, sticky="w")
        self.date_cd_var = tk.StringVar(value="HR")
        ttk.Entry(fixed_frame, textvariable=self.date_cd_var, width=8, state="readonly").grid(row=0, column=3, sticky="w")
        row += 1

        ttk.Label(main_frame, text="4. 조회 기간 (날짜 / 시간)", anchor="w").grid(row=row, column=0, sticky="w", pady=(10, 0))
        row += 1

        date_frame = ttk.Frame(main_frame)
        date_frame.grid(row=row, column=0, columnspan=4, sticky="w")
        row += 1

        years = [str(y) for y in range(2010, 2031)]
        months = [f"{m:02d}" for m in range(1, 13)]
        days = [f"{d:02d}" for d in range(1, 32)]
        hours = [f"{h:02d}" for h in range(0, 24)]

        ttk.Label(date_frame, text="시작일 (startDt)").grid(row=0, column=0, sticky="w")
        self.start_year_var = tk.StringVar(value=str(datetime.now().year))
        self.start_month_var = tk.StringVar(value=f"{datetime.now().month:02d}")
        self.start_day_var = tk.StringVar(value=f"{datetime.now().day:02d}")
        ttk.Combobox(date_frame, textvariable=self.start_year_var, values=years, width=6, state="readonly").grid(row=0, column=1, padx=(5, 0))
        ttk.Label(date_frame, text="년").grid(row=0, column=2)
        ttk.Combobox(date_frame, textvariable=self.start_month_var, values=months, width=4, state="readonly").grid(row=0, column=3)
        ttk.Label(date_frame, text="월").grid(row=0, column=4)
        ttk.Combobox(date_frame, textvariable=self.start_day_var, values=days, width=4, state="readonly").grid(row=0, column=5)
        ttk.Label(date_frame, text="일").grid(row=0, column=6, padx=(0, 10))
        ttk.Label(date_frame, text="시작시 (startHh)").grid(row=0, column=7, sticky="w", padx=(10, 0))
        self.start_hh_var = tk.StringVar(value="00")
        ttk.Combobox(date_frame, textvariable=self.start_hh_var, values=hours, width=4, state="readonly").grid(row=0, column=8)
        ttk.Label(date_frame, text="시").grid(row=0, column=9)

        ttk.Label(date_frame, text="종료일 (endDt)").grid(row=1, column=0, sticky="w", pady=(5, 0))
        self.end_year_var = tk.StringVar(value=str(datetime.now().year))
        self.end_month_var = tk.StringVar(value=f"{datetime.now().month:02d}")
        self.end_day_var = tk.StringVar(value=f"{datetime.now().day:02d}")
        ttk.Combobox(date_frame, textvariable=self.end_year_var, values=years, width=6, state="readonly").grid(row=1, column=1, padx=(5, 0), pady=(5, 0))
        ttk.Label(date_frame, text="년").grid(row=1, column=2, pady=(5, 0))
        ttk.Combobox(date_frame, textvariable=self.end_month_var, values=months, width=4, state="readonly").grid(row=1, column=3, pady=(5, 0))
        ttk.Label(date_frame, text="월").grid(row=1, column=4, pady=(5, 0))
        ttk.Combobox(date_frame, textvariable=self.end_day_var, values=days, width=4, state="readonly").grid(row=1, column=5, pady=(5, 0))
        ttk.Label(date_frame, text="일").grid(row=1, column=6, padx=(0, 10), pady=(5, 0))
        ttk.Label(date_frame, text="종료시 (endHh)").grid(row=1, column=7, sticky="w", padx=(10, 0), pady=(5, 0))
        self.end_hh_var = tk.StringVar(value="23")
        ttk.Combobox(date_frame, textvariable=self.end_hh_var, values=hours, width=4, state="readonly").grid(row=1, column=8, pady=(5, 0))
        ttk.Label(date_frame, text="시").grid(row=1, column=9, pady=(5, 0))

        # 5. 지점 번호 (stnIds) - label and combobox tightly stacked
        ttk.Label(main_frame, text="5. 지점 번호 (stnIds)", anchor="w").grid(row=row, column=0, sticky="w", pady=(0, 0))
        row += 1
        stations = [
            "강릉 105","강진군 259","강화 201","거제 294","거창 284","경주시 283","고산 185",
            "고창 172","고창군 251","고흥 262","광양시 266","광주 156","구미 279","군산 140",
            "금산 238","김해시 253","남원 247","남해 295","대관령 100","대구 143","대전 133",
            "동두천 98","동해 106","목포 165","문경 273","밀양 288","백령도 102","보령 235",
            "보성군 258","보은 226","봉화 271","부산 159","부안 243","부여 236","북강릉 104",
            "북창원 255","북춘천 93","산청 289","상주 137","서귀포 189","서산 129","서울 108",
            "성산 188","세종 239","속초 90","수원 119","순창군 254","순천 174","안동 136",
            "양산시 257","양평 202","여수 168","영광군 252","영덕 277","영월 121","영주 272",
            "영천 281","완도 170","울릉도 115","울산 152","울진 130","원주 114","의령군 263",
            "의성 278","이천 203","인제 211","인천 112","임실 244","장수 248","장흥 260",
            "전주 146","정선군 217","정읍 245","제주 184","제천 221","진도군 268","진주 192",
            "창원 155","천안 232","철원 95","청송군 276","청주 131","추풍령 135","춘천 101",
            "충주 127","태백 216","통영 162","파주 99","포항 138","함양군 264","합천 285",
            "해남 261","홍성 177","홍천 212","흑산도 169"
        ]
        self.stn_ids_var = tk.StringVar(value="수원 119")
        stn_combo = ttk.Combobox(main_frame, textvariable=self.stn_ids_var, values=stations, width=30)
        stn_combo.grid(row=row, column=0, sticky="w", pady=(0, 0))
        stn_combo.configure(state='normal')

        # Add simple autocomplete: filter station list on key release
        def _on_stn_keyrelease(event):
            typed = self.stn_ids_var.get().strip()
            if typed == "":
                stn_combo['values'] = stations
                return
            lowered = typed.lower()
            matches = [s for s in stations if lowered in s.lower()]
            # update dropdown values
            stn_combo['values'] = matches
            try:
                # open dropdown if there are matches
                if matches:
                    stn_combo.event_generate('<Down>')
            except Exception:
                pass

        stn_combo.bind('<KeyRelease>', _on_stn_keyrelease)

        # numOfRows removed from UI; fixed to 100 internally
        self.page_no_var = tk.StringVar(value="1")

        # 6. 생성된 요청 URL
        row += 1
        ttk.Label(main_frame, text="6. 생성된 요청 URL", anchor="w").grid(row=row, column=0, sticky="w", pady=(10, 0))
        row += 1
        self.url_text = tk.Text(main_frame, height=5, width=80, wrap="word")
        self.url_text.grid(row=row, column=0, columnspan=4, sticky="we")
        self.url_text.config(state="disabled")
        row += 1

        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=row, column=0, columnspan=4, pady=(10, 0), sticky="e")
        self.generate_btn = ttk.Button(btn_frame, text="URL 생성", command=self.generate_url)
        self.generate_btn.grid(row=0, column=0, padx=5)
        self.open_btn = ttk.Button(btn_frame, text="브라우저로 URL 열기", command=self.open_url_in_browser)
        self.open_btn.grid(row=0, column=1, padx=5)
        self.fetch_btn = ttk.Button(btn_frame, text="조회", command=self.fetch_and_show)
        self.fetch_btn.grid(row=0, column=2, padx=5)
        self.cancel_btn = ttk.Button(btn_frame, text="취소", command=self._cancel_fetch)
        self.cancel_btn.grid(row=0, column=3, padx=5)
        self.cancel_btn.state(['disabled'])
        # Button positions
        col = 4
        self.excel_btn = ttk.Button(btn_frame, text="엑셀로 저장", command=self.export_to_excel)
        self.excel_btn.grid(row=0, column=col, padx=5)
        col += 1
        self.cloud_upload_btn = ttk.Button(btn_frame, text="클라우드 업로드", command=self.cloud_upload_data)
        self.cloud_upload_btn.grid(row=0, column=col, padx=5)
        col += 1
        self.cloud_browser_btn = ttk.Button(btn_frame, text="클라우드 브라우저", command=self.cloud_browser)
        self.cloud_browser_btn.grid(row=0, column=col, padx=5)
        col += 1
        self.quit_btn = ttk.Button(btn_frame, text="종료", command=self.destroy)
        self.quit_btn.grid(row=0, column=col, padx=5)
        row += 1

        self.status_var = tk.StringVar(value="진행: 0%")
        self.progress = ttk.Progressbar(main_frame, orient='horizontal', length=200, mode='determinate', maximum=100)
        self.progress.grid(row=row, column=0, columnspan=3, sticky='we', pady=(6, 0))
        self.status_label = ttk.Label(main_frame, textvariable=self.status_var)
        self.status_label.grid(row=row, column=3, sticky='e', pady=(6, 0))
        row += 1

        tree_frame = ttk.Frame(main_frame)
        tree_frame.grid(row=row, column=0, columnspan=4, sticky="nsew", pady=(10, 0))
        main_frame.rowconfigure(row, weight=1)
        self.tree_frame = tree_frame
        self.result_tree = ttk.Treeview(tree_frame, show='headings')
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.result_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.result_tree.xview)
        self.result_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.result_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        self.bind('<Configure>', lambda e: self._adjust_tree_columns())
        self._cancel_event = threading.Event()
        self.summary_text = tk.Text(summary_frame, height=10, width=80, wrap='word', font=self.summary_font)
        self.summary_text.pack(fill=tk.BOTH, expand=True)

    def fetch_and_show(self):
        # Collect and validate inputs on the main thread, then spawn a background worker
        try:
            sy = int(self.start_year_var.get())
            sm = int(self.start_month_var.get())
            sd = int(self.start_day_var.get())
            sh = int(self.start_hh_var.get())
            ey = int(self.end_year_var.get())
            em = int(self.end_month_var.get())
            ed = int(self.end_day_var.get())
            eh = int(self.end_hh_var.get())
            start_dt = datetime(sy, sm, sd, sh)
            end_dt = datetime(ey, em, ed, eh)
        except Exception:
            messagebox.showerror("오류", "시작/종료일 또는 시간이 올바르지 않습니다.")
            return

        if end_dt < start_dt:
            messagebox.showerror("오류", "종료일시는 시작일시보다 이전일 수 없습니다.")
            return

        hours_between = int((end_dt - start_dt).total_seconds() // 3600)
        total_hours = hours_between + 2

        # enforce numOfRows=100 and compute last_page so numOfRows * last_page >= total_hours
        try:
            num_rows = 100
            import math
            last_page = max(1, math.ceil(total_hours / num_rows))
            # numOfRows is fixed to 100 and not editable by the user
            self.page_no_var.set(str(last_page))
        except Exception:
            last_page = int(self.page_no_var.get().strip() or "1")

        # Gather other parameters (read once on main thread)
        raw_key = self.service_key_var.get().strip()
        if not raw_key:
            messagebox.showerror("오류", "인증키(serviceKey)를 입력해주세요.")
            return

        data_type = self.data_type_var.get().upper()
        data_cd = self.data_cd_var.get()
        date_cd = self.date_cd_var.get()
        stn_input = self.stn_ids_var.get().strip()
        m = re.search(r"(\d+)", stn_input)
        stn_ids = m.group(1) if m else stn_input

        # disable fetch button during operation and enable cancel
        self.fetch_btn.state(['disabled'])
        try:
            self.cancel_btn.state(['!disabled'])
        except Exception:
            pass

        # reset progress and cancellation
        self._update_progress(0)
        self._cancel_event.clear()

        # start background worker
        worker_args = {
            'serviceKey': raw_key,
            'dataType': data_type,
            'dataCd': data_cd,
            'dateCd': date_cd,
            'startDt': f"{self.start_year_var.get()}{self.start_month_var.get()}{self.start_day_var.get()}",
            'startHh': self.start_hh_var.get(),
            'endDt': f"{self.end_year_var.get()}{self.end_month_var.get()}{self.end_day_var.get()}",
            'endHh': self.end_hh_var.get(),
            'stnIds': stn_ids,
            'numOfRows': '100',
        }

        t = threading.Thread(target=self._fetch_pages_worker, args=(worker_args, last_page, data_type), daemon=True)
        t.start()

    def _adjust_tree_columns(self):
        """Adjust Treeview column widths to fill available space in tree_frame."""
        try:
            cols = list(self.result_tree['columns'])
            if not cols:
                return
            # ensure geometry info is available
            width = self.tree_frame.winfo_width()
            if not width or width < 50:
                return
            # leave some room for vertical scrollbar
            scrollbar_space = 24
            available = max(100, width - scrollbar_space)
            per = max(60, int(available / len(cols)))
            for c in cols:
                self.result_tree.column(c, width=per)
        except Exception:
            # be silent on resize errors
            pass

    def _update_progress(self, percent):
        try:
            self.progress['value'] = percent
            self.status_var.set(f"진행: {percent}%")
        except Exception:
            pass

    def _populate_summary(self, items_sorted):
        """Generate a short summary (max ta, max hm, max enthalpy) and write to summary_text."""
        if not hasattr(self, 'summary_text'):
            return

        # helper to build datetime string from item
        def item_dt(it):
            y = it.get('year') or ''
            m = it.get('month') or ''
            d = it.get('day') or ''
            hh = it.get('hour') or ''
            parts = []
            if y and m and d:
                s = f"{y}-{m}-{d}"
                if hh:
                    s = f"{s} {hh}:00"
                return s
            # fallback: tm field
            if it.get('tm'):
                return it.get('tm')
            for k in ('time', 'date'):
                if it.get(k):
                    return it.get(k)
            return ''

        # find max 건구온도 (ta)
        def to_float(v):
            try:
                if v is None:
                    return None
                s = str(v).strip()
                if s == '':
                    return None
                return float(s.replace(',', ''))
            except Exception:
                return None

        max_ta = None
        max_ta_item = None
        max_hm = None
        max_hm_item = None
        max_ent = None
        max_ent_item = None

        for it in items_sorted:
            if not isinstance(it, dict):
                continue
            ta_v = to_float(it.get('ta'))
            hm_v = to_float(it.get('hm'))
            ent_v = None
            ekj = it.get('enthalpy_kj')
            if ekj:
                try:
                    ent_v = float(str(ekj))
                except Exception:
                    ent_v = None

            if ta_v is not None:
                if max_ta is None or ta_v > max_ta:
                    max_ta = ta_v
                    max_ta_item = it

            if hm_v is not None:
                if max_hm is None or hm_v > max_hm:
                    max_hm = hm_v
                    max_hm_item = it

            if ent_v is not None:
                if max_ent is None or ent_v > max_ent:
                    max_ent = ent_v
                    max_ent_item = it

        lines = []
        if max_ta_item:
            lines.append(f"최대 건구온도: {max_ta:.3f} ℃  (일시: {item_dt(max_ta_item)})")
        else:
            lines.append("최대 건구온도: 데이터 없음")

        if max_hm_item:
            lines.append(f"최대 상대습도: {max_hm:.3f} %RH  (일시: {item_dt(max_hm_item)})")
        else:
            lines.append("최대 상대습도: 데이터 없음")

        if max_ent_item:
            # show enthalpy with datetime and also the ta/hm on that same row
            ta_ent = to_float(max_ent_item.get('ta'))
            hm_ent = to_float(max_ent_item.get('hm'))
            lines.append(f"최대 엔탈피: {max_ent:.3f} kJ/kg  (일시: {item_dt(max_ent_item)})  ta={ta_ent if ta_ent is not None else ''} ℃, hm={hm_ent if hm_ent is not None else ''} %RH")
            # also show kcal
            try:
                kcal = max_ent / 4.1868
                lines.append(f" (동일 행 엔탈피: {kcal:.3f} kcal/kg)")
            except Exception:
                pass
        else:
            lines.append("최대 엔탈피: 데이터 없음")

        # TAC calculations (엔탈피 기준): top X% highest enthalpy, cut = smallest enthalpy within that top set
        try:
            import math
            ent_list = []
            ent_map = []  # list of (ent_val, item)
            for it in items_sorted:
                if not isinstance(it, dict):
                    continue
                ek = it.get('enthalpy_kj')
                try:
                    ev = float(str(ek))
                except Exception:
                    ev = None
                if ev is not None:
                    ent_list.append(ev)
                    ent_map.append((ev, it))

            if ent_list:
                ent_sorted_desc = sorted(ent_list, reverse=True)
                N = len(ent_sorted_desc)
                for pct in (1.0, 2.5):
                    top_n = max(1, math.ceil(N * pct / 100.0))
                    cutoff = ent_sorted_desc[top_n - 1]
                    # find earliest item in time order whose enthalpy equals cutoff (within tolerance)
                    match_item = None
                    for it in items_sorted:
                        if not isinstance(it, dict):
                            continue
                        try:
                            ev = float(str(it.get('enthalpy_kj') or ''))
                        except Exception:
                            continue
                        if abs(ev - cutoff) < 1e-9 or ev >= cutoff:
                            match_item = it
                            break

                    if match_item:
                        ta_v = to_float(match_item.get('ta'))
                        hm_v = to_float(match_item.get('hm'))
                        lines.append(f"TAC {pct}%(엔탈피기준): 컷오프 {cutoff:.3f} kJ/kg  (일시: {item_dt(match_item)})  ta={ta_v if ta_v is not None else ''} ℃, hm={hm_v if hm_v is not None else ''} %RH")
                    else:
                        lines.append(f"TAC {pct}%(엔탈피기준): 데이터 없음")

            # TAC calculations (온도 기준): for top X% highest 온도, report the lowest temperature within that top set and its datetime
            temp_list = []
            for it in items_sorted:
                if not isinstance(it, dict):
                    continue
                ta_v = to_float(it.get('ta'))
                if ta_v is not None:
                    temp_list.append(ta_v)

            if temp_list:
                temp_sorted_desc = sorted(temp_list, reverse=True)
                M = len(temp_sorted_desc)
                for pct in (1.0, 2.5):
                    top_m = max(1, math.ceil(M * pct / 100.0))
                    # the cutoff temp is the lowest temp in the top set (i.e., temp_sorted_desc[top_m-1])
                    cutoff_temp = temp_sorted_desc[top_m - 1]
                    # within items_sorted, find earliest item whose ta is >= cutoff_temp and collect those in the top set
                    top_items = []
                    for it in items_sorted:
                        if not isinstance(it, dict):
                            continue
                        try:
                            tv = to_float(it.get('ta'))
                        except Exception:
                            tv = None
                        if tv is None:
                            continue
                        if tv >= cutoff_temp:
                            top_items.append((tv, it))

                    if top_items:
                        # find the minimum temperature among the top_items
                        min_temp = min(t for t, _ in top_items)
                        # find earliest item whose temperature equals min_temp
                        match_item = None
                        for tval, it in top_items:
                            if abs(tval - min_temp) < 1e-9 or tval == min_temp:
                                match_item = it
                                break
                        if match_item:
                            lines.append(f"TAC {pct}%(온도기준): 상위 {pct}% 내 최소 온도 {min_temp:.3f} ℃  (일시: {item_dt(match_item)})")
                        else:
                            lines.append(f"TAC {pct}%(온도기준): 데이터 없음")
                    else:
                        lines.append(f"TAC {pct}%(온도기준): 데이터 없음")
        except Exception:
            pass

        # write to summary_text
        try:
            self.summary_text.config(state='normal')
            self.summary_text.delete('1.0', tk.END)
            self.summary_text.insert(tk.END, '\n'.join(lines))
            self.summary_text.config(state='disabled')
        except Exception:
            pass

    def _cancel_fetch(self):
        # Signal the background worker to stop after current page and disable cancel button
        try:
            if getattr(self, '_cancel_event', None):
                self._cancel_event.set()
            self.cancel_btn.state(['disabled'])
            self.status_var.set("취소 요청됨...")
        except Exception:
            pass

    def _on_fetch_complete(self, items):
        # called on main thread after worker finished; populate Treeview
        # First: parse any time-like field (prefer 'tm') into year/month/day/hour
        time_field_candidates = ('tm',)
        for it in items:
            if not isinstance(it, dict):
                continue
            raw = None
            for k in time_field_candidates:
                if k in it and it.get(k):
                    raw = it.get(k)
                    break
            if raw is None:
                # fallback: find any key that looks like time/date
                for k, v in it.items():
                    if v and re.search(r'time|tm|date|yy|mm|dd|hh', k, re.IGNORECASE):
                        raw = v
                        break

            y = mth = d = hh = ""
            if raw:
                parsed = None
                for fmt in ("%Y%m%d%H", "%Y%m%d%H%M", "%Y-%m-%d %H:%M", "%Y-%m-%d %H", "%Y%m%d", "%Y-%m-%d"):
                    try:
                        parsed = datetime.strptime(raw, fmt)
                        break
                    except Exception:
                        continue

                if parsed:
                    y = str(parsed.year)
                    mth = f"{parsed.month:02d}"
                    d = f"{parsed.day:02d}"
                    hh = f"{parsed.hour:02d}"
                else:
                    # try digit regex like YYYYMMDDHH
                    m = re.match(r"^(\d{4})[-/]?(\d{2})[-/]?(\d{2})(?:[ _T-]?(\d{2}))?", raw)
                    if m:
                        y = m.group(1)
                        mth = m.group(2)
                        d = m.group(3)
                        hh = m.group(4) or ""

            # attach parsed fields
            it['year'] = y
            it['month'] = mth
            it['day'] = d
            it['hour'] = hh
        # compute enthalpy columns based on ta (°C) and hm (%)
        def to_float(v):
            try:
                if v is None:
                    return None
                s = str(v).strip()
                if s == '':
                    return None
                s = s.replace(',', '').replace('\u200b', '')
                return float(s)
            except Exception:
                return None

        for it in items:
            if not isinstance(it, dict):
                continue
            ta_val = to_float(it.get('ta'))
            hm_val = to_float(it.get('hm'))
            ent_kj = None
            ent_kcal = None
            if ta_val is not None and hm_val is not None:
                try:
                    import math
                    # saturation vapor pressure (kPa) using Tetens formula
                    def sat_vp(Tc):
                        return 0.61078 * math.exp(17.27 * Tc / (Tc + 237.3))

                    Psat = sat_vp(ta_val)  # kPa
                    Pv = max(0.0, (hm_val / 100.0) * Psat)
                    P = 101.325  # kPa (approx atmospheric)
                    # humidity ratio W (kg water / kg dry air)
                    if P - Pv > 0:
                        W = 0.622 * Pv / (P - Pv)
                    else:
                        W = 0.0
                    # enthalpy (kJ/kg dry air)
                    ent_kj_val = 1.006 * ta_val + W * (2501 + 1.86 * ta_val)
                    ent_kj = round(ent_kj_val, 3)
                    ent_kcal = round(ent_kj_val / 4.1868, 3)
                except Exception:
                    ent_kj = None
                    ent_kcal = None

            it['enthalpy_kj'] = (f"{ent_kj:.3f}" if ent_kj is not None else "")
            it['enthalpy_kcal'] = (f"{ent_kcal:.3f}" if ent_kcal is not None else "")

        # desired display order: 지역번호, 지역명, 지점번호, 번호, 연도, 월, 일, 시간, 건구온도, 상대습도, 이슬점온도, 엔탈피(kcal/kg), 엔탈피(kJ/kg)
        desired_order = ['stnld', 'stnNm', 'stnIds', 'rnum', 'year', 'month', 'day', 'hour', 'ta', 'hm', 'td', 'enthalpy_kcal', 'enthalpy_kj']
        heading_map = {
            'stnld': '지역번호',
            'stnNm': '지역명',
            'stnIds': '지점번호',
            'rnum': '번호',
            'year': '연도',
            'month': '월',
            'day': '일',
            'hour': '시간',
            'ta': '건구온도 (℃DB)',
            'hm': '상대습도 (%RH)',
            'td': '이슬점온도',
            'enthalpy_kcal': '엔탈피 (kcal/kg)',
            'enthalpy_kj': '엔탈피 (kJ/kg)',
        }

        present_keys = set()
        for it in items:
            if isinstance(it, dict):
                present_keys.update(it.keys())

        cols = [k for k in desired_order if k in present_keys]
        if not cols:
            cols = []
            col_set = set()
            for it in items:
                if isinstance(it, dict):
                    for k in it.keys():
                        if k not in col_set:
                            col_set.add(k)
                            cols.append(k)

        # determine time key for sorting (prefer newly created 'year'+'month'+'day'+'hour')
        def sort_key(item):
            if not isinstance(item, dict):
                return 0
            if item.get('year') and item.get('month') and item.get('day'):
                try:
                    hh_val = item.get('hour') or '00'
                    s = f"{item.get('year')}{item.get('month')}{item.get('day')}{hh_val}"
                    return datetime.strptime(s, "%Y%m%d%H")
                except Exception:
                    pass
            # fallback: try tm or other time-like fields
            for k, v in item.items():
                if v and re.search(r'time|tm|date|yy|mm|dd|hh', k, re.IGNORECASE):
                    for fmt in ("%Y%m%d%H", "%Y-%m-%d %H:%M", "%Y%m%d", "%H%M", "%H"):
                        try:
                            return datetime.strptime(v, fmt)
                        except Exception:
                            continue
                    return v
            return 0

        try:
            items_sorted = sorted(items, key=sort_key)
        except Exception:
            items_sorted = items

        for c in self.result_tree.get_children():
            self.result_tree.delete(c)
        self.result_tree.config(columns=cols)
        for c in cols:
            self.result_tree.heading(c, text=heading_map.get(c, c))
            self.result_tree.column(c, width=120, anchor='w')

        for it in items_sorted:
            row_vals = [it.get(c, "") if isinstance(it, dict) else "" for c in cols]
            self.result_tree.insert('', tk.END, values=row_vals)

        # populate summary tab based on sorted items
        try:
            self._populate_summary(items_sorted)
        except Exception:
            pass

        self._adjust_tree_columns()
        self._update_progress(100)
        self.fetch_btn.state(['!disabled'])
        try:
            self.cancel_btn.state(['disabled'])
        except Exception:
            pass

    def _fetch_pages_worker(self, base_params, last_page, data_type):
        # Runs in background thread. Use self.after to perform UI updates on main thread.
        items = []

        def find_first_list(obj):
            if isinstance(obj, list):
                if obj and isinstance(obj[0], dict):
                    return obj
                for el in obj:
                    res = find_first_list(el)
                    if res:
                        return res
            elif isinstance(obj, dict):
                for v in obj.values():
                    res = find_first_list(v)
                    if res:
                        return res
            return None

        for p in range(1, last_page + 1):
            # check cancellation before each page
            if getattr(self, '_cancel_event', None) and self._cancel_event.is_set():
                # schedule partial update and stop
                self.after(0, lambda items=items: self._on_fetch_complete(items))
                return
            params = base_params.copy()
            params['pageNo'] = str(p)
            params['numOfRows'] = base_params.get('numOfRows', '100')
            query_str = "&".join(f"{k}={v}" for k, v in params.items())
            url = f"{BASE_URL}?{query_str}"

            try:
                resp = requests.get(url, timeout=15)
                resp.raise_for_status()
            except Exception as e:
                # schedule error message on main thread and continue
                self.after(0, lambda e=e, p=p: messagebox.showerror("요청 오류", f"페이지 {p} 요청 실패: {e}"))
                continue

            if data_type == 'JSON':
                try:
                    parsed = resp.json()
                except Exception as e:
                    self.after(0, lambda e=e, p=p: messagebox.showerror("파싱 오류", f"페이지 {p} JSON 파싱 실패: {e}"))
                    continue

                page_items = find_first_list(parsed) or []
                for it in page_items:
                    if isinstance(it, dict):
                        items.append(it)
            else:
                try:
                    root = ET.fromstring(resp.content)
                except Exception as e:
                    self.after(0, lambda e=e, p=p: messagebox.showerror("파싱 오류", f"페이지 {p} XML 파싱 실패: {e}"))
                    continue

                xml_items = root.findall('.//item')
                if not xml_items:
                    xml_items = [elem for elem in root.iter() if list(elem)]

                for it in xml_items:
                    d = {}
                    for child in list(it):
                        d[child.tag] = child.text
                    if d:
                        items.append(d)

            percent = int((p / last_page) * 100)
            self.after(0, lambda percent=percent: self._update_progress(percent))

            # check cancellation after page
            if getattr(self, '_cancel_event', None) and self._cancel_event.is_set():
                self.after(0, lambda items=items: self._on_fetch_complete(items))
                return

        # finished; schedule Treeview update on main thread
        self.after(0, lambda items=items: self._on_fetch_complete(items))

    def generate_url(self):
        raw_key = self.service_key_var.get().strip()
        if not raw_key:
            messagebox.showerror("오류", "인증키(serviceKey)를 입력해주세요.")
            return
        encoded_key = raw_key

        start_dt = f"{self.start_year_var.get()}{self.start_month_var.get()}{self.start_day_var.get()}"
        end_dt = f"{self.end_year_var.get()}{self.end_month_var.get()}{self.end_day_var.get()}"
        start_hh = self.start_hh_var.get()
        end_hh = self.end_hh_var.get()

        if not (len(start_dt) == 8 and start_dt.isdigit()):
            messagebox.showerror("오류", "시작일(startDt)이 올바르지 않습니다.")
            return
        if not (len(end_dt) == 8 and end_dt.isdigit()):
            messagebox.showerror("오류", "종료일(endDt)이 올바르지 않습니다.")
            return

        data_type = self.data_type_var.get()
        data_cd = self.data_cd_var.get()
        date_cd = self.date_cd_var.get()
        stn_input = self.stn_ids_var.get().strip()
        m = re.search(r"(\d+)", stn_input)
        stn_ids = m.group(1) if m else stn_input

        # numOfRows is fixed to 100 and not editable by the user
        num_rows = "100"
        page_no = self.page_no_var.get().strip() or "1"

        if not stn_ids:
            messagebox.showerror("오류", "지점 번호(stnIds)를 입력해주세요.")
            return

        params = {
            "serviceKey": encoded_key,
            "numOfRows": num_rows,
            "pageNo": page_no,
            "dataType": data_type,
            "dataCd": data_cd,
            "dateCd": date_cd,
            "startDt": start_dt,
            "startHh": start_hh,
            "endDt": end_dt,
            "endHh": end_hh,
            "stnIds": stn_ids,
        }

        query_str = "&".join(f"{k}={v}" for k, v in params.items())
        full_url = f"{BASE_URL}?{query_str}"

        self.url_text.config(state="normal")
        self.url_text.delete("1.0", tk.END)
        self.url_text.insert(tk.END, full_url)
        self.url_text.config(state="disabled")

    def open_url_in_browser(self):
        self.generate_url()
        self.url_text.config(state="normal")
        url = self.url_text.get("1.0", tk.END).strip()
        self.url_text.config(state="disabled")

        if url:
            webbrowser.open(url)
        else:
            messagebox.showwarning("경고", "먼저 'URL 생성'을 눌러 URL을 만들어주세요.")

    def export_to_excel(self):
        # Export the current Treeview contents to an XLSX file using openpyxl
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except Exception:
            messagebox.showerror("패키지 필요", "openpyxl이 필요합니다. 설치: pip install openpyxl")
            return

        cols = self.result_tree['columns']
        if not cols:
            messagebox.showinfo("정보", "저장할 표 내용이 없습니다.")
            return

        file_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel files', '*.xlsx')])
        if not file_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Results'

        # heading map to use Korean labels where available
        heading_map = {
            'stnld': '지역번호',
            'stnNm': '지역명',
            'stnIds': '지점번호',
            'rnum': '번호',
            'year': '연도',
            'month': '월',
            'day': '일',
            'hour': '시간',
            'tm': '일시',
            'ta': '건구온도 (℃DB)',
            'hm': '상대습도 (%RH)',
            'td': '이슬점온도',
            'enthalpy_kcal': '엔탈피 (kcal/kg)',
            'enthalpy_kj': '엔탈피 (kJ/kg)',
        }

        # write header (use Korean labels when possible)
        for i, c in enumerate(cols, start=1):
            ws.cell(row=1, column=i, value=heading_map.get(c, c))

        # write rows
        for r_idx, item_id in enumerate(self.result_tree.get_children(), start=2):
            vals = self.result_tree.item(item_id, 'values')
            for c_idx, v in enumerate(vals, start=1):
                ws.cell(row=r_idx, column=c_idx, value=v)

        # auto-width
        for i, c in enumerate(cols, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 20
        # write summary sheet if available
        try:
            if hasattr(self, 'summary_text'):
                summary = ''
                try:
                    summary = self.summary_text.get('1.0', tk.END).strip()
                except Exception:
                    summary = ''
                if summary:
                    ws_sum = wb.create_sheet(title='Summary')
                    for r, line in enumerate(summary.splitlines(), start=1):
                        ws_sum.cell(row=r, column=1, value=line)
                    # make column wider for readability
                    try:
                        ws_sum.column_dimensions['A'].width = 80
                    except Exception:
                        pass

            try:
                wb.save(file_path)
            except Exception as e:
                messagebox.showerror("저장 실패", str(e))
                return

            messagebox.showinfo("완료", f"엑셀 파일로 저장되었습니다: {file_path}")
        except Exception as e:
            messagebox.showerror("저장 실패", str(e))
            return
    
    def cloud_upload_data(self):
        """현재 데이터를 클라우드에 업로드"""
        cols = self.result_tree['columns']
        if not cols:
            messagebox.showinfo("정보", "업로드할 데이터가 없습니다. 먼저 조회를 실행하세요.")
            return
        
        # 임시 엑셀 파일 생성 (보안: 고유한 임시 파일 생성)
        import tempfile
        import os
        
        temp_fd, temp_file = tempfile.mkstemp(suffix=".xlsx", prefix="weather_")
        
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except Exception:
            messagebox.showerror("패키지 필요", "openpyxl이 필요합니다. 설치: pip install openpyxl")
            return
        
        try:
            # Close the file descriptor before writing with openpyxl
            os.close(temp_fd)
            
            # 임시 파일에 저장
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Weather Data'
            
            heading_map = {
                'stnld': '지역번호',
                'stnNm': '지역명',
                'stnIds': '지점번호',
                'rnum': '번호',
                'year': '연도',
                'month': '월',
                'day': '일',
                'hour': '시간',
                'tm': '일시',
                'ta': '건구온도 (℃DB)',
                'hm': '상대습도 (%RH)',
                'td': '이슬점온도',
                'enthalpy_kcal': '엔탈피 (kcal/kg)',
                'enthalpy_kj': '엔탈피 (kJ/kg)',
            }
            
            # 헤더 작성
            for i, c in enumerate(cols, start=1):
                ws.cell(row=1, column=i, value=heading_map.get(c, c))
            
            # 데이터 작성
            for r_idx, item_id in enumerate(self.result_tree.get_children(), start=2):
                vals = self.result_tree.item(item_id, 'values')
                for c_idx, v in enumerate(vals, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=v)
            
            wb.save(temp_file)
            
            # 클라우드 업로드 다이얼로그 열기
            from cloud_ui import CloudUploadDialog
            dialog = CloudUploadDialog(self, file_path=temp_file, data_type="weather_data")
            self.wait_window(dialog)
            
            # 임시 파일 삭제
            try:
                os.remove(temp_file)
            except Exception:
                pass
                
        except Exception as e:
            messagebox.showerror("오류", f"클라우드 업로드 준비 중 오류가 발생했습니다:\n{e}")
    
    def cloud_browser(self):
        """클라우드 브라우저 열기"""
        try:
            from cloud_ui import CloudBrowserDialog
            dialog = CloudBrowserDialog(self)
            self.wait_window(dialog)
        except Exception as e:
            messagebox.showerror("오류", f"클라우드 브라우저를 여는 중 오류가 발생했습니다:\n{e}")


if __name__ == "__main__":
    try:
        app = AsosGUI()
        app.mainloop()
    except Exception:
        import traceback, os
        tb = traceback.format_exc()
        log_path = os.path.join(os.path.dirname(__file__), 'asos_gui_error.log')
        try:
            with open(log_path, 'w', encoding='utf-8') as f:
                f.write(tb)
        except Exception:
            pass
        # also print to console so developer can see immediate traceback
        print(tb)
        # re-raise to allow environment to show the error as well
        raise
